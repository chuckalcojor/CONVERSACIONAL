import argparse
import os
import re
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_RELACION_CLIENTES = Path(r"C:\Users\gasto\Desktop\DASHBORAD\ZIDONG-main\Informacion\Relacion Clientes.xlsx")


def _nit_candidates(value) -> list[str]:
    raw = str(value or "").strip()
    clean = re.sub(r"[^0-9]", "", raw)
    candidates = []

    def add(item: str) -> None:
        if item and item not in candidates:
            candidates.append(item)

    add(raw)
    add(clean)
    if len(clean) > 1:
        add(clean[:-1])
        add(f"{clean[:-1]}-{clean[-1]}")
    return candidates


def _find_client_by_tax_id(client, tax_id) -> dict | None:
    for nit in _nit_candidates(tax_id):
        rows = client.table("clients").select("id, clinic_name, tax_id").eq("tax_id", nit).limit(1).execute().data or []
        if rows:
            return rows[0]
    return None


def _load_env(path: str | None) -> None:
    load_dotenv(path or ROOT / ".env")


def main() -> int:
    parser = argparse.ArgumentParser(description="Carga asignaciones cliente -> motorizado desde Relacion Clientes.xlsx.")
    parser.add_argument("--env-file", help="Ruta a .env con SUPABASE_URL y SUPABASE_SERVICE_ROLE_KEY")
    parser.add_argument("--source", default=str(DEFAULT_RELACION_CLIENTES), help="Ruta al Excel Relacion Clientes.xlsx")
    args = parser.parse_args()

    _load_env(args.env_file)
    client = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_ROLE_KEY"])
    couriers = {
        row["name"].strip().lower(): row
        for row in client.table("couriers").select("id, name").execute().data or []
        if row.get("name")
    }

    ws = load_workbook(args.source, read_only=True, data_only=True)["Hoja1"]
    upserts = []
    skipped = 0
    unmatched = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        _codigo, nombre, tax_id, _address, _poblacion, zona, mensajero = row[:7]
        if not mensajero or not isinstance(zona, (int, float)):
            skipped += 1
            continue
        courier = couriers.get(str(mensajero).strip().lower())
        found = _find_client_by_tax_id(client, tax_id)
        if not courier or not found:
            unmatched.append({"clinic_name": nombre, "tax_id": tax_id, "zone": zona, "courier": mensajero})
            continue
        upserts.append({
            "client_id": found["id"],
            "courier_id": courier["id"],
            "assigned_by": "seed:relacion_clientes_zonas",
        })

    if upserts:
        deduped = {row["client_id"]: row for row in upserts}
        upserts = list(deduped.values())
        client.table("client_courier_assignment").upsert(upserts, on_conflict="client_id").execute()

    print(f"assignments_upserted={len(upserts)}")
    print(f"rows_skipped_no_zone_or_courier={skipped}")
    print(f"rows_unmatched={len(unmatched)}")
    for item in unmatched[:20]:
        print("unmatched", item)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
