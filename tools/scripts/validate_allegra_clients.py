from collections import Counter, defaultdict
from pathlib import Path
import re
import sys

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from app.config import SUPABASE_SERVICE_ROLE_KEY, SUPABASE_URL


EXCEL_PATH = Path(r"C:\Users\gasto\Desktop\Alegra - Terceros.xlsx")


def clean_text(value) -> str:
    return str(value or "").strip()


def normalize_key(value: str) -> str:
    text = clean_text(value).lower().translate(str.maketrans("áéíóúüñ", "aeiouun"))
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def normalize_phone(*values) -> str:
    for value in values:
        digits = re.sub(r"\D+", "", clean_text(value))
        if digits:
            return digits
    return ""


def parse_name(raw_name: str) -> tuple[str, str | None, str]:
    raw_name = clean_text(raw_name)
    match = re.match(r"^\(([^)]+)\)\s*(.*)$", raw_name)
    if match:
        return match.group(1).strip(), match.group(2).strip() or None, "parenthesized_clinic"
    if ")" in raw_name and "(" not in raw_name:
        clinic_name, professional_name = raw_name.split(")", 1)
        return clinic_name.strip(), professional_name.strip() or None, "missing_open_parenthesis"
    return raw_name, None, "direct_name"


def load_excel_rows() -> list[dict]:
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    records = []
    for row_number, values in enumerate(rows, start=2):
        row = dict(zip(headers, values))
        clinic_name, professional_name, pattern = parse_name(row.get("Nombre"))
        invoice_status = clean_text(row.get("Factura electrónica"))
        records.append({
            "row_number": row_number,
            "external_code": clean_text(row.get("Identificación")),
            "clinic_name": clinic_name,
            "clinic_key": normalize_key(clinic_name),
            "professional_name": professional_name,
            "professional_key": normalize_key(professional_name or ""),
            "source_pattern": pattern,
            "tax_id": clean_text(row.get("Identificación")),
            "phone": normalize_phone(row.get("Celular"), row.get("Teléfono 1"), row.get("Teléfono 2")),
            "address": clean_text(row.get("Dirección")),
            "city": clean_text(row.get("Municipio")),
            "department": clean_text(row.get("Departamento")),
            "email": clean_text(row.get("Correo")),
            "invoice_status": invoice_status,
            "electronic_invoicing": invoice_status.upper() == "SI",
            "is_deleted": invoice_status.lower() == "eliminado",
            "raw_name": clean_text(row.get("Nombre")),
        })
    return records


def load_db_clients() -> list[dict]:
    db = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
    return db.table("clients").select("id, clinic_name, tax_id, phone, address, city, is_active").limit(5000).execute().data or []


def build_index(rows: list[dict], field: str) -> dict[str, list[dict]]:
    index = defaultdict(list)
    for row in rows:
        value = normalize_key(row.get(field) or "") if field == "clinic_name" else clean_text(row.get(field))
        if value:
            index[value].append(row)
    return index


def print_group(title: str, rows: list[dict], limit: int = 20) -> None:
    print(f"\n{title}: {len(rows)}")
    for row in rows[:limit]:
        status = " [ELIMINADO]" if row["is_deleted"] else ""
        nit = row["tax_id"] if row["tax_id"] else "SIN NIT"
        print(f"  fila={row['row_number']} nit={nit} clinica={row['clinic_name']}{status} profesional={row['professional_name'] or '-'}")
    if len(rows) > limit:
        print(f"  ... {len(rows) - limit} mas")


def main() -> None:
    load_dotenv()
    excel_rows = load_excel_rows()
    all_rows = excel_rows
    active_rows = [row for row in excel_rows if not row["is_deleted"]]
    deleted_rows = [row for row in excel_rows if row["is_deleted"]]
    db_clients = load_db_clients()

    db_by_tax = build_index(db_clients, "tax_id")
    db_by_name = build_index(db_clients, "clinic_name")

    existing_by_tax = [row for row in all_rows if row["tax_id"] and row["tax_id"] in db_by_tax]
    existing_by_name = [row for row in all_rows if row["clinic_key"] in db_by_name]

    deleted_existing_by_name = [row for row in deleted_rows if row["clinic_key"] in db_by_name]
    deleted_new = [row for row in deleted_rows if row["clinic_key"] not in db_by_name]

    new_active_rows = [
        row for row in active_rows
        if not (row["tax_id"] and row["tax_id"] in db_by_tax) and row["clinic_key"] not in db_by_name
    ]
    no_nit_rows = [row for row in all_rows if not row["tax_id"] or not row["tax_id"].strip()]

    all_tax_groups = defaultdict(list)
    for row in all_rows:
        if row["tax_id"] and not row["is_deleted"]:
            all_tax_groups[row["tax_id"]].append(row)
    sucursal_groups = {tax: branches for tax, branches in all_tax_groups.items() if len(branches) > 1}

    duplicated_tax = [tax for tax, count in Counter(row["tax_id"] for row in active_rows if row["tax_id"]).items() if count > 1]
    duplicated_name = [name for name, count in Counter(row["clinic_key"] for row in active_rows if row["clinic_key"]).items() if count > 1]
    missing_required = [row for row in all_rows if not row["clinic_name"] or not row["address"]]
    malformed_parenthesis = [
        row for row in all_rows
        if ")" in row["raw_name"] and row["source_pattern"] != "parenthesized_clinic"
    ]

    total_to_insert = len(new_active_rows) + len(deleted_new)
    total_to_update = len(existing_by_name)

    print("VALIDACION ALEGRA -> PLATAFORMA (incluye eliminados)")
    print(f"archivo={EXCEL_PATH}")
    print(f"total_excel={len(excel_rows)}")
    print(f"activos={len(active_rows)}")
    print(f"eliminados={len(deleted_rows)} (se incluiran con is_active=False)")
    print(f"sin_nit={len(no_nit_rows)} (se incluiran con tax_id=None)")
    print(f"con_veterinaria_entre_parentesis={sum(1 for row in all_rows if row['source_pattern'] == 'parenthesized_clinic')}")
    print(f"con_profesional_responsable={sum(1 for row in all_rows if row['professional_name'])}")
    print(f"ya_existen_por_identificacion={len(existing_by_tax)}")
    print(f"ya_existen_por_nombre={len(existing_by_name)}")
    print(f"nuevos_activos_para_insertar={len(new_active_rows)}")
    print(f"eliminados_existentes_actualizar={len(deleted_existing_by_name)}")
    print(f"eliminados_nuevos_insertar={len(deleted_new)}")
    print(f"total_a_cargar={total_to_insert + total_to_update}")
    print(f"  nuevos_insertar={total_to_insert}")
    print(f"  existentes_actualizar={total_to_update}")
    print(f"sucursales_identificadas={len(sucursal_groups)} grupos")
    print(f"identificaciones_duplicadas_en_excel={len(duplicated_tax)}")
    print(f"nombres_duplicados_en_excel={len(duplicated_name)}")
    print(f"campos_nit_faltantes={len(no_nit_rows)} (se cargan sin NIT)")
    print(f"campos_direccion_faltantes={len([r for r in all_rows if not r['address']])}")
    print(f"nombres_con_parentesis_malformado={len(malformed_parenthesis)}")

    print_group("nuevos_activos", new_active_rows)
    print_group("eliminados_nuevos", deleted_new)
    print_group("sin_nit", no_nit_rows)
    print_group("parentesis_malformado", malformed_parenthesis)

    if sucursal_groups:
        print("\nSucursales (mismo NIT, distintas sedes):")
        for tax, branches in sucursal_groups.items():
            print(f"  NIT {tax} ({len(branches)} sedes):")
            for b in branches:
                addr = f" - {b['address']}" if b['address'] else ""
                print(f"    {b['clinic_name']}{addr}")

    if duplicated_name:
        print("\nnombres_duplicados_muestra:")
        for name in duplicated_name[:20]:
            rows = [row for row in all_rows if row["clinic_key"] == name]
            print(f"  clinica={rows[0]['clinic_name']} -> nits={[row['tax_id'] for row in rows]}")


if __name__ == "__main__":
    main()