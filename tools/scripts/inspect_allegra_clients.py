from pathlib import Path
import re
from collections import Counter, defaultdict

from openpyxl import load_workbook


EXCEL_PATH = Path(r"C:\Users\gasto\Desktop\Alegra - Terceros.xlsx")


def _text(value) -> str:
    return str(value or "").strip()


def _key(value: str) -> str:
    text = value.lower().translate(str.maketrans("áéíóúüñ", "aeiouun"))
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def parse_name(raw_name: str) -> dict:
    raw_name = _text(raw_name)
    match = re.match(r"^\(([^)]+)\)\s*(.*)$", raw_name)
    if match:
        clinic_name = match.group(1).strip()
        professional_name = match.group(2).strip() or None
        return {
            "raw_name": raw_name,
            "clinic_name": clinic_name,
            "professional_name": professional_name,
            "source_pattern": "parenthesized_clinic",
        }

    if ")" in raw_name and "(" not in raw_name:
        clinic_name, professional_name = raw_name.split(")", 1)
        return {
            "raw_name": raw_name,
            "clinic_name": clinic_name.strip(),
            "professional_name": professional_name.strip() or None,
            "source_pattern": "missing_open_parenthesis",
        }

    return {
        "raw_name": raw_name,
        "clinic_name": raw_name,
        "professional_name": None,
        "source_pattern": "direct_name",
    }


def _branch_base(clinic_name: str) -> str:
    text = clinic_name
    text = re.sub(r"\b(sede|suc\.?|sucursal|principal)\b", " ", text, flags=re.I)
    text = re.sub(r"\b(norte|sur|centro|occidente|oriente|suba|kennedy|bosa|soacha|chia|timiza|modelia|fontibon|engativa|usaquen|chapinero|toberin|cedritos|restrepo|galerias|castilla|venecia|madelena)\b", " ", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip(" -_")
    return text or clinic_name


def load_rows() -> list[dict]:
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() for value in next(rows)]
    records = []
    for row_number, values in enumerate(rows, start=2):
        record = dict(zip(headers, values))
        parsed = parse_name(record.get("Nombre"))
        phone = _text(record.get("Celular")) or _text(record.get("Teléfono 1")) or _text(record.get("Teléfono 2"))
        invoice_status = _text(record.get("Factura electrónica"))
        clinic_name = parsed["clinic_name"]
        records.append({
            "row_number": row_number,
            "clinic_key": _key(clinic_name),
            "clinic_name": clinic_name,
            "professional_name": parsed["professional_name"],
            "source_pattern": parsed["source_pattern"],
            "tax_id": _text(record.get("Identificación")),
            "department": _text(record.get("Departamento")),
            "city": _text(record.get("Municipio")),
            "address": _text(record.get("Dirección")),
            "phone": phone,
            "email": _text(record.get("Correo")),
            "invoice_status": invoice_status,
            "is_deleted": invoice_status.lower() == "eliminado",
            "branch_base_key": _key(_branch_base(clinic_name)),
            "branch_base_name": _branch_base(clinic_name),
            "raw_name": parsed["raw_name"],
        })
    return records


def print_summary(records: list[dict]) -> None:
    active = [row for row in records if not row["is_deleted"]]
    print(f"total_registros={len(records)}")
    print(f"activos_o_no_eliminados={len(active)}")
    print(f"eliminados={len(records) - len(active)}")
    print("factura_electronica=", dict(Counter(row["invoice_status"] for row in records)))
    print("patrones_nombre=", dict(Counter(row["source_pattern"] for row in records)))
    print("sin_direccion=", sum(1 for row in active if not row["address"]))
    print("sin_telefono=", sum(1 for row in active if not row["phone"]))
    print("sin_identificacion=", sum(1 for row in active if not row["tax_id"]))
    print("sin_correo=", sum(1 for row in active if not row["email"]))

    duplicate_tax_ids = [tax_id for tax_id, count in Counter(row["tax_id"] for row in active if row["tax_id"]).items() if count > 1]
    print(f"identificaciones_duplicadas={len(duplicate_tax_ids)}")
    if duplicate_tax_ids:
        print("primeras_identificaciones_duplicadas=", duplicate_tax_ids[:20])

    grouped = defaultdict(list)
    for row in active:
        grouped[row["branch_base_key"]].append(row)
    branches = {key: rows for key, rows in grouped.items() if len(rows) > 1}
    print(f"posibles_grupos_con_sucursales={len(branches)}")
    for rows in list(branches.values())[:25]:
        names = [row["clinic_name"] for row in rows]
        print(f"sucursal_base={rows[0]['branch_base_name']} cantidad={len(rows)} nombres={names}")

    print("\nejemplos_parseados:")
    for row in records[:30]:
        print({
            "fila": row["row_number"],
            "clinica": row["clinic_name"],
            "profesional": row["professional_name"],
            "estado": row["invoice_status"],
            "ciudad": row["city"],
        })


def main() -> None:
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"archivo={EXCEL_PATH}")
    print(f"hojas={workbook.sheetnames}")

    for sheet in workbook.worksheets:
        print(f"\n[{sheet.title}] filas={sheet.max_row} columnas={sheet.max_column}")
        for index, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 6), values_only=True),
            start=1,
        ):
            print(f"{index}: {row}")

    print("\n[resumen_normalizacion]")
    print_summary(load_rows())


if __name__ == "__main__":
    main()
